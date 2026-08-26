"""
Automação de cobrança de extrato bancário via Digisac (WhatsApp)
==================================================================

O que este script faz:
1. Fica rodando continuamente e verifica a planilha de clientes
   (colunas: NOME, CPF, NUMERO DE TELEFONE, STATUS DE ENVIO DO EXTRATO)
   a cada X segundos (INTERVALO_VERIFICACAO_SEGUNDOS no config.py) —
   essa verificação acontece SEMPRE, esteja a planilha alterada ou não
2. Filtra quem está com status "PENDENTE"
3. Para cada pendente, verifica no log (log_envios.csv) quando foi o último
   envio para aquele telefone:
   - Se nunca enviou, ou já passou a JANELA_REENVIO_HORAS (config.py) -> envia
   - Se enviou há menos tempo que isso -> NÃO envia de novo, só avisa no
     terminal que viu o pendente, mas que ainda está dentro do prazo de espera
4. Registra cada envio (ou tentativa) no log_envios.csv, SEM apagar o
   histórico anterior (log é usado como memória entre execuções)

ANTES DE USAR:
- Preencha as variáveis em config.py (subdomínio, token, ID da conexão de WhatsApp)
- Ajuste INTERVALO_VERIFICACAO_SEGUNDOS e JANELA_REENVIO_HORAS no config.py
  conforme sua necessidade
- Rode primeiro em modo DRY_RUN=True para conferir os destinatários e o texto,
  sem enviar nada de verdade
- Depois rode com DRY_RUN=False para enviar de verdade
- O script fica rodando para sempre. Para parar, use Ctrl+C no terminal.

Como rodar:
    pip install requests openpyxl --break-system-packages
    python3 cobranca_extrato.py
"""

import csv
import os
import sys
import time
from datetime import datetime

import requests
from openpyxl import load_workbook

import config


# ---------------------------------------------------------------------------
# Configurações fixas do script (não precisa mexer aqui)
# ---------------------------------------------------------------------------

BASE_URL = f"https://{config.SUBDOMAIN}.{config.DOMINIO}/api/v1"
HEADERS = {
    "Authorization": f"Bearer {config.API_TOKEN}",
    "Content-Type": "application/json",
}

STATUS_PENDENTE = "PENDENTE"

# Colunas esperadas na planilha (ajuste aqui se sua planilha usar outros nomes)
COL_NOME = "NOME"
COL_TELEFONE = "NUMERO DE TELEFONE"
COL_STATUS = "STATUS DE ENVIO DO EXTRATO"

LOG_PATH = "log_envios.csv"

# Valores padrão caso não existam no config.py
INTERVALO_VERIFICACAO_SEGUNDOS = getattr(
    config, "INTERVALO_VERIFICACAO_SEGUNDOS", 300)
JANELA_REENVIO_HORAS = getattr(config, "JANELA_REENVIO_HORAS", 1)


def limpar_telefone(numero: str) -> str:
    """Remove formatação e garante o padrão internacional (55 + DDD + número)."""
    digitos = "".join(c for c in str(numero) if c.isdigit())
    # remove eventual zero à esquerda do DDD
    if digitos.startswith("0"):
        digitos = digitos[1:]
    if not digitos.startswith("55"):
        digitos = "55" + digitos
    return digitos


def ler_pendentes(caminho_planilha: str):
    """Lê a planilha e retorna a lista de clientes com status PENDENTE."""
    wb = load_workbook(caminho_planilha, data_only=True)
    ws = wb.active

    # mapeia nome da coluna -> índice, a partir do cabeçalho (linha 1)
    cabecalho = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {nome: i for i, nome in enumerate(cabecalho)}

    faltando = [c for c in (COL_NOME, COL_TELEFONE, COL_STATUS) if c not in idx]
    if faltando:
        print(f"ERRO: não encontrei as colunas {faltando} na planilha.")
        print(f"Colunas encontradas: {cabecalho}")
        sys.exit(1)

    pendentes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx[COL_NOME]] is None:
            continue
        status = str(row[idx[COL_STATUS]] or "").strip().upper()
        if status == STATUS_PENDENTE:
            pendentes.append({
                "nome": str(row[idx[COL_NOME]]).strip(),
                "telefone": limpar_telefone(row[idx[COL_TELEFONE]]),
            })
    return pendentes


def _levantar_com_detalhe(resp: requests.Response):
    """Levanta erro incluindo o corpo da resposta da API (essencial pra debugar)."""
    if not resp.ok:
        raise RuntimeError(
            f"HTTP {resp.status_code} em {resp.request.method} {resp.url}\n"
            f"Corpo da resposta: {resp.text[:1000]}"
        )


def buscar_ou_criar_contato(telefone: str, nome: str) -> str:
    """Cria o contato no Digisac (ou reaproveita se o número já existir). Retorna o contactId."""
    resp = requests.post(
        f"{BASE_URL}/contacts",
        headers=HEADERS,
        json={"name": nome, "number": telefone, "serviceId": config.CONNECTION_ID},
        timeout=30,
    )
    if resp.status_code in (200, 201):
        return resp.json()["id"]

    # Se já existir um contato com esse número, a API costuma avisar no corpo.
    # Tentamos extrair o id do próprio erro, se ele vier junto.
    try:
        corpo = resp.json()
    except ValueError:
        corpo = {}
    contato_existente = corpo.get("id") or (corpo.get("contact") or {}).get("id")
    if contato_existente:
        return contato_existente

    _levantar_com_detalhe(resp)


def enviar_mensagem(contact_id: str, texto: str):
    """Envia a mensagem de texto para o contato via Digisac."""
    resp = requests.post(
        f"{BASE_URL}/messages",
        headers=HEADERS,
        json={
            "contactId": contact_id,
            "serviceId": config.CONNECTION_ID,
            "type": "chat",
            "text": texto,
        },
        timeout=30,
    )
    _levantar_com_detalhe(resp)
    return resp.json()


def calcular_mes_competencia(referencia: datetime = None) -> str:
    """
    Retorna o mês/ano de competência dos documentos no formato MM/AAAA,
    sempre referente ao mês ANTERIOR ao mês em que a mensagem está sendo enviada.
    Ex.: se a mensagem é enviada em setembro/2026, retorna "08/2026".
    """
    referencia = referencia or datetime.now()
    ano = referencia.year
    mes = referencia.month - 1
    if mes == 0:
        mes = 12
        ano -= 1
    return f"{mes:02d}/{ano}"


def montar_mensagem(nome: str) -> str:
    nome_completo = nome.strip().title()
    mes_competencia = calcular_mes_competencia()
    return config.MENSAGEM_TEMPLATE.format(nome=nome_completo, mes_competencia=mes_competencia)


# ---------------------------------------------------------------------------
# Log como memória entre execuções (novo)
# ---------------------------------------------------------------------------

def carregar_ultimo_envio_por_telefone(log_path: str) -> dict:
    """
    Lê o log_envios.csv já existente e retorna um dicionário:
        { telefone: datetime do último envio bem-sucedido (ENVIADO ou SIMULADO) }

    Isso é o que permite ao script "lembrar", mesmo depois de reiniciado,
    quando cada cliente recebeu a última mensagem.
    """
    ultimo_envio = {}
    if not os.path.exists(log_path):
        return ultimo_envio

    with open(log_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("status") not in ("ENVIADO", "SIMULADO"):
                continue
            try:
                dt = datetime.fromisoformat(row["data_hora"])
            except (ValueError, KeyError):
                continue
            telefone = row.get("telefone")
            if not telefone:
                continue
            anterior = ultimo_envio.get(telefone)
            if anterior is None or dt > anterior:
                ultimo_envio[telefone] = dt
    return ultimo_envio


def registrar_log(log_path: str, nome: str, telefone: str, status: str, detalhe: str):
    """Acrescenta uma linha ao log, SEM apagar o que já existia (modo append)."""
    arquivo_novo = not os.path.exists(log_path)
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if arquivo_novo:
            writer.writerow(["data_hora", "nome", "telefone", "status", "detalhe"])
        writer.writerow([datetime.now().isoformat(), nome, telefone, status, detalhe])


# ---------------------------------------------------------------------------
# Processamento de uma "rodada" (quando o arquivo muda)
# ---------------------------------------------------------------------------

def processar_planilha():
    pendentes = ler_pendentes(config.PLANILHA)
    print(f"  -> {len(pendentes)} cliente(s) com status PENDENTE encontrados na planilha.")

    # Carrega o histórico de envios a cada rodada, pois o log pode ter
    # sido atualizado por uma execução anterior ou por outra instância.
    ultimo_envio = carregar_ultimo_envio_por_telefone(LOG_PATH)

    for cliente in pendentes:
        nome = cliente["nome"]
        telefone = cliente["telefone"]

        envio_anterior = ultimo_envio.get(telefone)
        if envio_anterior is not None:
            horas_desde_envio = (datetime.now() - envio_anterior).total_seconds() / 3600
            if horas_desde_envio < JANELA_REENVIO_HORAS:
                faltam = JANELA_REENVIO_HORAS - horas_desde_envio
                print(
                    f"  [AGUARDANDO] {nome} ({telefone}): já recebeu mensagem em "
                    f"{envio_anterior:%d/%m/%Y %H:%M} (há {horas_desde_envio:.1f}h). "
                    f"Ainda dentro do prazo de {JANELA_REENVIO_HORAS}h "
                    f"-> faltam {faltam:.1f}h para poder reenviar."
                )
                continue  # não envia, pula para o próximo cliente

        mensagem = montar_mensagem(nome)

        if config.DRY_RUN:
            print(f"  [SIMULAÇÃO] Enviaria para {nome} ({telefone}):\n{mensagem}\n{'-'*50}")
            registrar_log(LOG_PATH, nome, telefone, "SIMULADO", "-")
            ultimo_envio[telefone] = datetime.now()
            continue

        try:
            contact_id = buscar_ou_criar_contato(telefone, nome)
            enviar_mensagem(contact_id, mensagem)
            print(f"  [OK] Mensagem enviada para {nome} ({telefone})")
            registrar_log(LOG_PATH, nome, telefone, "ENVIADO", "-")
            ultimo_envio[telefone] = datetime.now()
        except Exception as e:
            print(f"  [ERRO] Falha ao enviar para {nome} ({telefone}): {e}")
            registrar_log(LOG_PATH, nome, telefone, "ERRO", str(e))

        time.sleep(config.INTERVALO_ENTRE_ENVIOS_SEGUNDOS)


# ---------------------------------------------------------------------------
# Loop principal: roda para sempre, só reage quando o arquivo muda
# ---------------------------------------------------------------------------

def main():
    print("Automação de cobrança iniciada.")
    print(f"Planilha monitorada: {config.PLANILHA}")
    print(f"Verificando pendentes a cada {INTERVALO_VERIFICACAO_SEGUNDOS}s | "
          f"Janela de reenvio: {JANELA_REENVIO_HORAS}h")
    print("Pressione Ctrl+C para parar.\n")

    while True:
        print(f"[{datetime.now():%d/%m/%Y %H:%M:%S}] Verificando planilha...")
        try:
            processar_planilha()
        except FileNotFoundError:
            print(f"[ERRO] Planilha não encontrada em: {config.PLANILHA}")
        print(f"[{datetime.now():%d/%m/%Y %H:%M:%S}] Verificação concluída. "
              f"Próxima em {INTERVALO_VERIFICACAO_SEGUNDOS}s.\n")

        time.sleep(INTERVALO_VERIFICACAO_SEGUNDOS)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nEncerrado pelo usuário.")